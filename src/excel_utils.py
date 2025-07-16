import pandas as pd
from openpyxl import load_workbook


def carregar_planilha(caminho):
    dfs = pd.read_excel(caminho, sheet_name=None, engine="openpyxl")
    return dfs["XYZ"], dfs["Gestores"], dfs["Logs"]


def salvar_logs_e_atualizacoes(
    caminho, df_colaboradores, df_log, aba_analise=None, df_analise=None
):
    from openpyxl.utils.dataframe import dataframe_to_rows

    book = load_workbook(caminho)

    # Atualiza a aba 'XYZ'
    if "XYZ" in book.sheetnames:
        sheet = book["XYZ"]
        sheet.delete_rows(2, sheet.max_row)  # Mantém o cabeçalho
        for row in dataframe_to_rows(df_colaboradores, index=False, header=False):
            sheet.append(row)

    # Atualiza ou cria a aba 'Logs'
    if "Logs" not in book.sheetnames:
        book.create_sheet("Logs")
    sheet_logs = book["Logs"]
    start_row = sheet_logs.max_row + 1
    for _, row in df_log.iterrows():
        for col_idx, value in enumerate(row, 1):
            sheet_logs.cell(row=start_row, column=col_idx, value=value)
        start_row += 1

    # Atualiza ou cria aba de análise (opcional)
    if aba_analise and df_analise is not None:
        if aba_analise not in book.sheetnames:
            book.create_sheet(aba_analise)
        else:
            sheet = book[aba_analise]
            sheet.delete_rows(2, sheet.max_row)
        sheet = book[aba_analise]
        for row in dataframe_to_rows(df_analise, index=False, header=False):
            sheet.append(row)

    # Salva alterações
    book.save(caminho)
